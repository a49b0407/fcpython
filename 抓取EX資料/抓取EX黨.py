# -*- coding: utf-8 -*-
"""
Created on Thu Apr 15 21:24:49 2021

@author: user
"""
import openpyxl
f='sales.xlsx'
wb=openpyxl.load_workbook(f)
ws=wb.get_sheet_by_name('2020Q1')#棄用了
#  WB['2020q1']現代使用比較方便
print(ws['A1'].value)
print(ws['A2'].value)
print(ws['B2'].value)
print(ws['B3'].value)
print(ws['C5'].value)
print(ws['E3'].column)
print(ws['E3'].row)
print(ws['E3'].coordinate)
