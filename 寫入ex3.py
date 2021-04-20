# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 18:30:06 2021

@author: user
"""
import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet('學生表',1)#建立
wb.create_sheet('出席紀錄',0)#建立
ws = wb['學生表']#建立
title=['姓名','科系']#建立
ws.append(title)
while True:
    stu=list()
    name=input('學生姓名:')
    if name == '':
        break
    dep=input('科系:')
    stu.append(name)
    stu.append(dep)
    ws.append(stu)
wb.save('demo2.xlsx')










